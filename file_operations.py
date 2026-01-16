import asyncio
import uuid
import datetime
import traceback
import os
from io import BytesIO
from fastapi import Body, File, HTTPException, UploadFile, Path
from fastapi.responses import JSONResponse, StreamingResponse
from PyPDF2 import PdfReader
from openpyxl import load_workbook
from gridfs import NoFile

# 导入全局应用实例和数据库集合
from main import app, col_file_meta, fs

# 导入dashscope用于摘要生成
from dashscope import Generation

# -------------------------- 文件操作工具函数 --------------------------
async def extract_file_content(file: UploadFile, file_content: bytes):
    """智能提取各类文件内容（升级后）：图片OCR/音频转文字/PDF/Word/Excel/TXT/视频帧等"""
    content = ""
    try:
        filename = file.filename.lower()
        if any(ext in filename for ext in [".jpg", ".jpeg", ".png", ".bmp", ".gif"]):
            # 图片：当前使用模拟内容（实际需接入OCR服务）
            content = f"[{file.filename}] 图片内容，已保存"
        elif any(ext in filename for ext in [".mp3", ".wav", ".m4a", ".ogg"]):
            # 音频：当前使用模拟内容（实际需接入语音转文字服务）
            content = f"[{file.filename}] 音频内容，已保存"
        elif any(ext in filename for ext in [".mp4", ".avi", ".mov", ".wmv"]):
            # 视频：当前使用模拟内容（实际需接入视频分析服务）
            content = f"[{file.filename}] 视频内容，已保存"
        elif ".pdf" in filename:
            # PDF：使用PyPDF2提取文字
            pdf_reader = PdfReader(BytesIO(file_content))
            for page in pdf_reader.pages:
                content += page.extract_text() or ""
            if not content.strip():
                content = f"[{file.filename}] PDF文件，已保存，无文字内容"
        elif ".docx" in filename:
            # Word：当前使用模拟内容（实际需接入python-docx库）
            content = f"[{file.filename}] Word文档，已保存"
        elif ".xlsx" in filename:
            # Excel：使用openpyxl提取内容
            workbook = load_workbook(filename=BytesIO(file_content))
            sheet = workbook.active
            for row in sheet.iter_rows(values_only=True):
                if any(cell for cell in row):  # 只处理非空行
                    content += "\t".join(str(cell) if cell is not None else "" for cell in row) + "\n"
        elif ".txt" in filename:
            # TXT：直接读取
            content = file_content.decode("utf-8", errors="ignore")
        else:
            # 其他格式：保存为二进制文件
            content = f"[{file.filename}] 已保存，不支持内容提取"
    except Exception as e:
        print(f"文件内容提取失败: {e}")
        content = f"[文件提取失败] {str(e)}"
    return content

# 使用dashscope生成文本摘要
def generate_summary(text: str, max_length: int = 500) -> str:
    """使用dashscope API生成文本摘要"""
    try:
        prompt = f"请提取以下文本的核心摘要，控制在{max_length}字以内：\n\n{text}"
        response = Generation.call(
            model='qwen-turbo',
            prompt=prompt,
            max_tokens=max_length,
            temperature=0.3,
            top_p=0.8
        )
        if response.status_code == 200 and response.output.choices:
            return response.output.choices[0].message.content.strip()
        else:
            print(f"摘要生成失败: {response}")
            return text[:max_length]  # 失败时返回原文前max_length个字符
    except Exception as e:
        print(f"摘要生成异常: {e}")
        return text[:max_length]  # 异常时返回原文前max_length个字符

# -------------------------- 文件操作API接口 --------------------------
@app.post("/save_all", summary="核心接口：发送任意内容/文件给AI，永久保存+自动整理")
async def save_everything(
    content: str = Body(default="", description="文本内容：日记/资料/文字信息"),
    files: list[UploadFile] = File(default=[], description="上传文件：图片/音频/视频/PDF/Excel/所有格式")
):
    user_id = "user_001"
    save_id = str(uuid.uuid4())
    create_time = datetime.datetime.now()
    
    try:
        if content.strip():
            # 使用dashscope生成摘要，替换之前的langextract.extract
            core_content = generate_summary(content, max_length=500)
            col_file_meta.insert_one({
                "save_id": save_id,
                "user_id": user_id,
                "type": "text",
                "content": content,
                "core_content": core_content,
                "filename": "文本笔记/日记",
                "create_time": create_time,
                "is_valid": True
            })
        
        file_list = []
        for file in files:
            try:
                # 只读取一次文件内容
                file_content = await file.read()
                file_id = fs.put(file_content, filename=file.filename, content_type=file.content_type)
                extract_content = await extract_file_content(file, file_content)
                col_file_meta.insert_one({
                    "save_id": save_id,
                    "user_id": user_id,
                    "type": "file",
                    "file_id": file_id,
                    "filename": file.filename,
                    "content": extract_content,
                    "create_time": create_time,
                    "is_valid": True
                })
                file_list.append(file.filename)
            except Exception as e:
                print(f"处理文件 {file.filename} 失败: {e}")
                return JSONResponse(status_code=500, content={"code": 500, "msg": f"处理文件 {file.filename} 失败: {str(e)}"})
        
        # 不再直接调用conversation_chain，避免循环导入
        # 更新记忆的功能可以通过其他方式实现，比如在聊天接口中处理
        
        return {
            "code": 200,
            "msg": "✅ 所有内容已永久保存+AI自动整理完成",
            "data": {"save_id": save_id, "create_time": str(create_time), "files": file_list}
        }
    except Exception as e:
        print(f"保存内容失败: {e}")
        return JSONResponse(status_code=500, content={"code": 500, "msg": f"保存内容失败: {str(e)}"})

@app.get("/download_file/{file_id}", summary="下载文件接口：根据文件ID下载文件")
async def download_file(file_id: str = Path(..., description="文件ID，从file_id字段获取")):
    try:
        # 查找文件元信息
        file_meta = col_file_meta.find_one({"file_id": file_id})
        if not file_meta:
            raise HTTPException(status_code=404, detail="文件不存在")
        
        # 从GridFS获取文件
        grid_out = fs.get(file_id)
        filename = grid_out.filename
        
        # 返回文件流
        return StreamingResponse(
            grid_out,
            media_type=grid_out.content_type or "application/octet-stream",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except NoFile:
        raise HTTPException(status_code=404, detail="文件不存在")
    except Exception as e:
        print(f"下载文件失败: {e}")
        return JSONResponse(status_code=500, content={"code": 500, "msg": f"下载文件失败: {str(e)}"})

@app.post("/list_files", summary="获取文件列表接口：列出所有已保存的文件，支持按类型过滤")
async def list_files(
    user_id: str = Body(default="user_001", description="用户ID"),
    page: int = Body(default=1, description="页码"),
    page_size: int = Body(default=10, description="每页数量"),
    file_type: str = Body(default=None, description="文件类型过滤，如pdf, docx等")
):
    try:
        # 计算跳过的文件数
        skip = (page - 1) * page_size
        
        # 构建查询条件
        query = {"user_id": user_id, "type": "file", "is_valid": True}
        
        # 如果提供了文件类型，添加文件名过滤条件
        if file_type:
            # 确保文件类型以点开头
            if not file_type.startswith("."):
                file_type = f".{file_type}"
            # 不区分大小写地匹配文件名
            query["filename"] = {"$regex": file_type, "$options": "i"}
        
        # 查询文件列表
        files = list(col_file_meta.find(
            query,
            {"_id": 0, "file_id": 1, "filename": 1, "content": 1, "create_time": 1}
        ).skip(skip).limit(page_size))
        
        # 获取总文件数
        total = col_file_meta.count_documents(query)
        
        return {
            "code": 200,
            "msg": "获取文件列表成功",
            "data": {
                "files": files,
                "page": page,
                "page_size": page_size,
                "total": total
            }
        }
    except Exception as e:
        print(f"获取文件列表失败: {e}")
        return JSONResponse(status_code=500, content={"code": 500, "msg": f"获取文件列表失败: {str(e)}"})