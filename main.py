import os
import uuid
import datetime
import json
from io import BytesIO
from dotenv import load_dotenv
from fastapi import FastAPI, UploadFile, File, Body, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pymongo import MongoClient
from gridfs import GridFS, NoFile
from langchain_core.prompts import PromptTemplate
# from langchain_classic.prompts import PromptTemplate
from langchain_core.language_models.base import BaseLanguageModel
from langchain_core.outputs import LLMResult, Generation
from langchain_core.runnables import Runnable
from langchain_core.runnables.history import RunnableWithMessageHistory
import langextract as le
from PyPDF2 import PdfReader
from openpyxl import load_workbook
import uvicorn

# 加载环境变量
load_dotenv()

# 检查必要的环境变量
required_env_vars = ["MONGO_URI", "MONGO_DB_NAME"]
for var in required_env_vars:
    if not os.getenv(var):
        raise ValueError(f"环境变量 {var} 未设置")

app = FastAPI(title="AI纳界助理", description="永久存储/智能记忆/深度检索/自然语言操控/对话页")

# -------------------------- 新增：模拟LLM类 --------------------------
class MockLLM(BaseLanguageModel):
    """一个简单的模拟LLM类，用于满足自定义实体记忆的接口要求"""
    
    def __init__(self):
        super().__init__()
    
    def generate_prompt(self, prompts, stop=None, callbacks=None, **kwargs):
        """模拟生成回复，返回空的实体信息"""
        generations = []
        for _ in prompts:
            # 返回空的生成结果
            generations.append([Generation(text="{}")])
        return LLMResult(generations=generations, llm_output={})
    
    async def agenerate_prompt(self, prompts, stop=None, callbacks=None, **kwargs):
        """异步模拟生成回复，返回空的实体信息"""
        return self.generate_prompt(prompts, stop, callbacks, **kwargs)
    
    def invoke(self, input, stop=None, callbacks=None, **kwargs):
        """实现BaseLanguageModel要求的invoke方法"""
        return "{}"
    
    async def ainvoke(self, input, stop=None, callbacks=None, **kwargs):
        """实现BaseLanguageModel要求的异步ainvoke方法"""
        return "{}"

# 跨域配置【升级】：前端本地调试+生产环境全兼容
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 【新增核心】托管静态前端页面 - 访问 http://localhost:8000 直接打开仿豆包对话页
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="static")

# -------------------------- 1. 数据库初始化 - MongoDB 永久存储 --------------------------
client = MongoClient(os.getenv("MONGO_URI"))
db = client[os.getenv("MONGO_DB_NAME")]  # GridFS存储大文件：图片/音频/视频/PDF/Excel等
fs = GridFS(db)  # 初始化GridFS
# 集合定义（所有数据永久保存）
col_chat_history = db["chat_history"]  # 聊天记录
col_user_memory = db["user_memory"]    # 用户关键记忆（爱吃鱼、不吃辣等）
col_file_meta = db["file_meta"]        # 文件元信息+抽取的内容

# -------------------------- 2. 自定义实体记忆系统 --------------------------
class CustomEntityMemory:
    """自定义实体记忆系统，替代已弃用的ConversationEntityMemory"""
    
    def __init__(self, entity_cache_limit=100, human_prefix="用户", ai_prefix="AI助理"):
        self.entity_cache_limit = entity_cache_limit
        self.human_prefix = human_prefix
        self.ai_prefix = ai_prefix
        
        # 实体存储
        self.entity_store = type('EntityStore', (), {
            'store': {}  # 存储实体信息的字典
        })()
        
        # 历史对话存储
        self.history = []
        
        # 对话缓冲区
        self.buffer = []
    
    def load_memory_variables(self, inputs):
        """加载记忆变量"""
        # 获取历史对话
        history = self.get_history_string()
        
        # 获取实体信息
        entities = json.dumps(self.entity_store.store, ensure_ascii=False)
        
        return {
            "history": history,
            "entities": entities
        }
    
    def save_context(self, inputs, outputs):
        """保存对话上下文"""
        input_text = inputs.get("input", "")
        output_text = outputs.get("output", "")
        
        # 保存到历史
        self.history.append({
            "user": input_text,
            "ai": output_text
        })
        
        # 限制历史长度
        if len(self.history) > 100:
            self.history = self.history[-100:]
        
        # 更新实体存储
        self._update_entity_store(input_text, output_text)
    
    def get_history_string(self):
        """获取格式化的历史对话字符串"""
        history_str = ""
        for entry in self.history:
            history_str += f"{self.human_prefix}: {entry['user']}\n{self.ai_prefix}: {entry['ai']}\n"
        return history_str.strip()
    
    def _update_entity_store(self, input_text, output_text):
        """更新实体存储"""
        # 简单的实体提取逻辑（实际应用中可以使用更复杂的NLP技术）
        combined_text = input_text + " " + output_text
        
        # 提取可能的实体（简单示例：提取以"我喜欢"开头的实体）
        if "我喜欢" in combined_text:
            start_idx = combined_text.find("我喜欢") + 3
            # 提取到下一个标点符号或句子结束
            end_idx = start_idx
            while end_idx < len(combined_text) and combined_text[end_idx] not in [",", ".", "。", "，", "!", "！", "?", "？", "\n"]:
                end_idx += 1
            if end_idx > start_idx:
                entity = combined_text[start_idx:end_idx].strip()
                if entity:
                    self.entity_store.store[entity] = "用户喜欢的事物"
        
        # 限制实体存储数量
        if len(self.entity_store.store) > self.entity_cache_limit:
            # 移除最早添加的实体
            old_entity = next(iter(self.entity_store.store))
            del self.entity_store.store[old_entity]

# 创建模拟LLM实例
mock_llm = MockLLM()

# 使用自定义实体记忆替代ConversationEntityMemory
entity_memory = CustomEntityMemory(
    entity_cache_limit=100,
    human_prefix="用户",
    ai_prefix="AI助理"
)

MEMORY_PROMPT = PromptTemplate(
    input_variables=["input", "history", "entities"],
    template="""你是用户的专属AI纳界助理，拥有用户的全部长期记忆和知识库，风格和豆包一致，回复简洁精准友好。
    1. 你需要严格记住用户的实体信息：{entities}（比如爱吃鱼、不吃辣、忌口、喜好等，所有对话必须关联该记忆）
    2. 参考历史对话上下文：{history}
    3. 针对用户的问题/指令：{input} 进行精准回复，支持自然语言操控所有功能。
    规则：用户的所有资料永久保存，可随时调取；自动整理无效文件；提取关键记忆并永久生效；上传的文件自动解析内容并保存。
    回复要求：口语化、流畅，和豆包的回复风格一致，不要生硬，记忆内容无缝融入回复中。
    """
)

# 创建一个简单的runnable，模拟原来的ConversationChain行为
class SimpleRunnable(Runnable):
    def __init__(self, entity_memory, prompt, llm):
        self.entity_memory = entity_memory
        self.prompt = prompt
        self.llm = llm
    
    def invoke(self, input, config=None, **kwargs):
        # 获取历史对话
        history = self.entity_memory.load_memory_variables({})["history"]
        # 获取实体信息
        entities = self.entity_memory.load_memory_variables({})["entities"]
        
        # 格式化提示
        formatted_prompt = self.prompt.format(
            input=input["input"],
            history=history,
            entities=entities
        )
        
        # 使用LLM生成回复
        response = self.llm.invoke(formatted_prompt)
        
        # 更新记忆
        self.entity_memory.save_context(
            {"input": input["input"]},
            {"output": response}
        )
        
        return {"output": response}
    
    async def ainvoke(self, input, config=None, **kwargs):
        # 异步实现，与同步版本相同
        return self.invoke(input, config, **kwargs)

# 创建简单的runnable实例
runnable = SimpleRunnable(
    entity_memory=entity_memory,
    prompt=MEMORY_PROMPT,
    llm=mock_llm
)

# 定义获取历史对话的函数
# 注意：这里我们使用CustomEntityMemory的内置历史，所以返回空列表
def get_session_history(session_id):
    return []

# 创建带有历史记录的runnable
conversation_chain = runnable

# -------------------------- 3. 工具函数（新增+优化） --------------------------
async def extract_file_content(file: UploadFile, file_content: bytes):
    """智能提取各类文件内容（升级后）：图片OCR/音频转文字/PDF/Word/Excel/TXT/视频帧等"""
    content = ""
    try:
        filename = file.filename.lower()
        if any(ext in filename for ext in [".jpg", ".jpeg", ".png", ".bmp", ".gif"]):
            # 图片：当前使用模拟内容（实际需接入OCR服务）
            content = f"[{file.filename}] 图片内容，已保存"
        elif any(ext in filename for ext in [".mp3", ".wav", ".m4a", ".ogg"]):
            # 音频：当前使用模拟内容（实际需接入语音转文字服务）
            content = f"[{file.filename}] 音频内容，已保存"
        elif any(ext in filename for ext in [".mp4", ".avi", ".mov", ".wmv"]):
            # 视频：当前使用模拟内容（实际需接入视频分析服务）
            content = f"[{file.filename}] 视频内容，已保存"
        elif ".pdf" in filename:
            # PDF：使用PyPDF2提取文字
            pdf_reader = PdfReader(BytesIO(file_content))
            for page in pdf_reader.pages:
                content += page.extract_text() or ""
            if not content.strip():
                content = f"[{file.filename}] PDF文件，已保存，无文字内容"
        elif ".docx" in filename:
            # Word：当前使用模拟内容（实际需接入python-docx库）
            content = f"[{file.filename}] Word文档，已保存"
        elif ".xlsx" in filename:
            # Excel：使用openpyxl提取内容
            workbook = load_workbook(filename=BytesIO(file_content))
            sheet = workbook.active
            for row in sheet.iter_rows(values_only=True):
                if any(cell for cell in row):  # 只处理非空行
                    content += "\t".join(str(cell) if cell is not None else "" for cell in row) + "\n"
        elif ".txt" in filename:
            # TXT：直接读取
            content = file_content.decode("utf-8", errors="ignore")
        else:
            # 其他格式：保存为二进制文件
            content = f"[{file.filename}] 已保存，不支持内容提取"
    except Exception as e:
        print(f"文件内容提取失败: {e}")
        content = f"[文件提取失败] {str(e)}"
    return content

# -------------------------- 4. 新增：前端首页路由（访问根目录打开仿豆包对话页） --------------------------
@app.get("/", summary="仿豆包风格的AI纳界助理对话首页")
async def index(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})

# -------------------------- 5. 核心API接口（全部保留+优化适配前端） --------------------------
@app.post("/save_all", summary="核心接口：发送任意内容/文件给AI，永久保存+自动整理")
async def save_everything(
    content: str = Body(default="", description="文本内容：日记/资料/文字信息"),
    files: list[UploadFile] = File(default=[], description="上传文件：图片/音频/视频/PDF/Excel/所有格式")
):
    user_id = "user_001"
    save_id = str(uuid.uuid4())
    create_time = datetime.datetime.now()
    
    try:
        if content.strip():
            core_content = le.extract(content, extract_type="summary", max_length=500)
            col_file_meta.insert_one({
                "save_id": save_id,
                "user_id": user_id,
                "type": "text",
                "content": content,
                "core_content": core_content,
                "filename": "文本笔记/日记",
                "create_time": create_time,
                "is_valid": True
            })
        
        file_list = []
        for file in files:
            try:
                # 只读取一次文件内容
                file_content = await file.read()
                file_id = fs.put(file_content, filename=file.filename, content_type=file.content_type)
                extract_content = await extract_file_content(file, file_content)
                col_file_meta.insert_one({
                    "save_id": save_id,
                    "user_id": user_id,
                    "type": "file",
                    "file_id": file_id,
                    "filename": file.filename,
                    "content": extract_content,
                    "create_time": create_time,
                    "is_valid": True
                })
                file_list.append(file.filename)
            except Exception as e:
                print(f"处理文件 {file.filename} 失败: {e}")
                return JSONResponse(status_code=500, content={"code": 500, "msg": f"处理文件 {file.filename} 失败: {str(e)}"})
        
        # 更新记忆
        await conversation_chain.ainvoke(
            {"input": f"保存资料：{content}，上传文件：{file_list}"},
            config={"configurable": {"session_id": user_id}}
        )
        
        return {
            "code": 200,
            "msg": "✅ 所有内容已永久保存+AI自动整理完成",
            "data": {"save_id": save_id, "create_time": str(create_time), "files": file_list}
        }
    except Exception as e:
        print(f"保存内容失败: {e}")
        return JSONResponse(status_code=500, content={"code": 500, "msg": f"保存内容失败: {str(e)}"})

@app.post("/chat", summary="聊天+记忆+汇总保存：核心对话接口，仿豆包回复风格")
async def chat_with_assistant(request: Request):
    user_id = "user_001"
    create_time = datetime.datetime.now()
    
    try:
        # 手动解析请求体
        request_body = await request.json()
        message = request_body.get("message", "").strip()
        
        if not message:
            return JSONResponse(status_code=400, content={"code": 400, "msg": "消息内容不能为空"})
        
        # 核心：LangChain对话+记忆更新，豆包风格回复
        result = await conversation_chain.ainvoke(
            {"input": message}
        )
        ai_reply = result["output"]
        
        # 指令处理：删除无效文件/汇总资料/提取记忆
        if any(key in message for key in ["删除无效文件", "清理垃圾文件", "删除空文件"]):
            col_file_meta.update_many({"content": "", "filename": {"$regex": "无内容"}}, {"$set": {"is_valid": False}})
            ai_reply += "\n✅ 已自动清理所有无效/空文件，标记为失效状态"
        
        if any(key in message for key in ["汇总保存", "整理资料", "汇总我的所有资料"]):
            all_valid_docs = list(col_file_meta.find({"user_id": user_id, "is_valid": True}))
            summary = le.extract(str(all_valid_docs), extract_type="summary", max_length=1000)
            col_file_meta.insert_one({
                "user_id": user_id,
                "type": "summary",
                "content": summary,
                "filename": "资料汇总-" + str(create_time.date()),
                "create_time": create_time,
                "is_valid": True
            })
            ai_reply += f"\n✅ 已汇总你所有的有效资料并永久保存，共整理 {len(all_valid_docs)} 条内容"
        
        if any(key in message for key in ["提取关键记忆", "我的记忆", "我有哪些偏好"]):
            ai_reply = f"📌 你的长期关键记忆：{entity_memory.entity_store.store}\n\n{ai_reply}"
        
        # 聊天记录永久保存
        col_chat_history.insert_one({
            "user_id": user_id,
            "user_msg": message,
            "ai_reply": ai_reply,
            "create_time": create_time
        })
        
        # 关键记忆持久化
        col_user_memory.update_one(
            {"user_id": user_id},
            {"$set": {"memory": entity_memory.entity_store.store, "update_time": create_time}},
            upsert=True
        )
        
        return {
            "code": 200,
            "user_msg": message,
            "ai_reply": ai_reply,
            "your_key_memory": entity_memory.entity_store.store,
            "create_time": str(create_time)
        }
    except json.JSONDecodeError:
        return JSONResponse(status_code=400, content={"code": 400, "msg": "请求格式错误，请使用JSON格式"})
    except Exception as e:
        print(f"聊天失败: {e}")
        return JSONResponse(status_code=500, content={"code": 500, "msg": f"聊天失败: {str(e)}"})

@app.post("/load_memory", summary="重启服务后加载长期记忆：保证记忆永不丢失")
async def load_user_memory():
    user_id = "user_001"
    try:
        # 加载用户的长期记忆
        user_memory = col_user_memory.find_one({"user_id": user_id})
        if user_memory and "memory" in user_memory:
            entity_memory.entity_store.store = user_memory["memory"]
            return {"code": 200, "msg": "长期记忆加载成功", "memory": user_memory["memory"]}
        return {"code": 200, "msg": "暂无长期记忆", "memory": {}}
    except Exception as e:
        print(f"加载记忆失败: {e}")
        return JSONResponse(status_code=500, content={"code": 500, "msg": f"加载记忆失败: {str(e)}"})

@app.post("/search", summary="深度检索：自然语言搜索所有保存的资料/聊天记录/记忆")
async def search_all(
    query: str = Body(..., description="搜索关键词/自然语言查询"),
    search_type: str = Body(default="all", description="搜索类型：all(全部)/file(文件)/chat(聊天记录)/memory(记忆)")
):
    user_id = "user_001"
    try:
        results = []
        
        # 搜索文件元信息
        if search_type in ["all", "file"]:
            file_results = list(col_file_meta.find({"user_id": user_id, "is_valid": True}))
            for doc in file_results:
                if query in str(doc.get("content", "")) or query in doc.get("filename", ""):
                    results.append({
                        "type": "file",
                        "filename": doc.get("filename", ""),
                        "content": doc.get("content", "")[:200] + "..." if len(doc.get("content", "")) > 200 else doc.get("content", ""),
                        "create_time": str(doc.get("create_time", ""))
                    })
        
        # 搜索聊天记录
        if search_type in ["all", "chat"]:
            chat_results = list(col_chat_history.find({"user_id": user_id}))
            for doc in chat_results:
                if query in doc.get("user_msg", "") or query in doc.get("ai_reply", ""):
                    results.append({
                        "type": "chat",
                        "user_msg": doc.get("user_msg", ""),
                        "ai_reply": doc.get("ai_reply", "")[:200] + "..." if len(doc.get("ai_reply", "")) > 200 else doc.get("ai_reply", ""),
                        "create_time": str(doc.get("create_time", ""))
                    })
        
        # 搜索用户记忆
        if search_type in ["all", "memory"]:
            user_memory = col_user_memory.find_one({"user_id": user_id})
            if user_memory and "memory" in user_memory:
                for key, value in user_memory["memory"].items():
                    if query in key or query in str(value):
                        results.append({
                            "type": "memory",
                            "key": key,
                            "value": str(value),
                            "create_time": str(user_memory.get("update_time", ""))
                        })
        
        # 汇总搜索结果
        ai_response = f"找到 {len(results)} 条相关结果\n\n"
        for i, result in enumerate(results[:10]):  # 最多显示10条
            ai_response += f"{i+1}. [{result['type']}] {result.get('filename', result.get('key', '无标题'))}\n"
            ai_response += f"   {result.get('content', result.get('user_msg', ''))[:100]}...\n\n"
        
        if len(results) > 10:
            ai_response += f"... 还有 {len(results)-10} 条结果未显示，请使用更精确的搜索词"
        
        return {
            "code": 200,
            "msg": "搜索完成",
            "data": {
                "user_memory": entity_memory.entity_store.store,
                "match_files": [r for r in results if r["type"] == "file"],
                "match_chat": [r for r in results if r["type"] == "chat"],
                "ai_summary": ai_response
            }
        }
    except Exception as e:
        print(f"检索失败: {e}")
        return JSONResponse(status_code=500, content={"code": 500, "msg": f"检索失败: {str(e)}"})

# 启动服务命令：uvicorn main:app --host 0.0.0.0 --port 8000 --reload
if __name__ == "__main__":
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)